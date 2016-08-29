#-------------------------------------------------------------------------------
#! -*- coding:utf-8 -*-
# Name:        模块2
# Purpose:
#
# Author:      zhx
#
# Created:     06/05/2016
# Copyright:   (c) zhx 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import openpyxl
import jieba
def main():
    cctv_data = openpyxl.load_workbook("cctv.xlsx")
    cctv_new = openpyxl.Workbook()
    new_sheet = cctv_new.active
    #new_sheet2 = cctv_new.active
    #print cctv_data.get_sheet_names()
    sheet = cctv_data["Sheet"]
#内容	标题长度	标题内容	评论	点赞	转发	链接	视频	图片	艾特	时间	话题	星期
    total_dict = {}
    for r in xrange(2,4748):
        print r
        for c in xrange(3,4):
            #tmp=sheet.cell(row=r,column=3).value#标题
            tmp=sheet.cell(row=r,column=1).value#内容
            good = sheet.cell(row=r,column=5).value
            if c==3 and tmp!=None:
                seg_list = jieba.cut(tmp)
                words = ' '.join(seg_list)
                words = words.split(' ')
                get_repeat_out=[]
                for w in words:
                    if w in get_repeat_out:
                        continue
                    get_repeat_out.append(w)
                    if len(w)==1:
                        continue
                    if w==' ':
                        continue
                    if total_dict.has_key(w):
                        total_dict[w].append(good)
                    else:
                        a = []
                        a.append(good)
                        total_dict[w] =a
    r =0
    for w in total_dict.keys():
        if r==0:
            r=1
            continue
        new_sheet.cell(row=r,column=1).value =w
        good_max = 0
        good_min = 99999
        good_mean = 0
        for num in total_dict[w]:
            if num!='赞':
                num = int(num)
                if num>good_max:
                    good_max = num
                if num<good_min:
                    good_min = num
                good_mean+=num
        good_mean = int (good_mean/len(total_dict[w]))
        new_sheet.cell(row=r,column=2).value =good_min
        new_sheet.cell(row=r,column=3).value =good_max
        new_sheet.cell(row=r,column=4).value =good_mean
        r+=1
    cctv_new.save("cctv_keywords.xlsx")



main()

