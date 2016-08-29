#-------------------------------------------------------------------------------
# Name:        濠碘槅鍨埀顒冩珪閸?
# Purpose:
#
# Author:      zhx
#
# Created:     17/05/2016
# Copyright:   (c) zhx 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import openpyxl

def main():
    cctv_data = openpyxl.load_workbook('train_all_feature_reg.xlsx')
    cctv_new = openpyxl.Workbook()
    new_sheet = cctv_new.active
    sheet = cctv_data["Sheet"]
    cnt = 2
    for r in xrange(1,4749):
        if r==1:
            for i in xrange(1,29):
                new_sheet.cell(row=1,column=i).value = sheet.cell(row=r,column=i).value
            continue
        print r
        picture = sheet.cell(row=r,column=4).value
        min_min = sheet.cell(row=r,column=9).value
        min_max = sheet.cell(row=r,column=10).value
        min_mean = sheet.cell(row=r,column=11).value
        num_posi = sheet.cell(row=r,column=18).value
        if num_posi == None:
            continue
        if min_min == None:
            continue
        elif min_min==99999:
            continue
        else:
            for i in range(1,29):
                if i ==4 :
                    new_sheet.cell(row=cnt,column=i).value = int (sheet.cell(row=r,column=i).value)
                else:
                    new_sheet.cell(row=cnt,column=i).value = sheet.cell(row=r,column=i).value
##            new_sheet.cell(row=cnt,column=2).value = sheet.cell(row=r,column=2).value
##            new_sheet.cell(row=cnt,column=3).value = sheet.cell(row=r,column=3).value
##            new_sheet.cell(row=cnt,column=4).value = int(sheet.cell(row=r,column=4).value)
##            new_sheet.cell(row=cnt,column=5).value = sheet.cell(row=r,column=5).value
##            new_sheet.cell(row=cnt,column=6).value = sheet.cell(row=r,column=6).value
##            new_sheet.cell(row=cnt,column=7).value = sheet.cell(row=r,column=7).value
##            new_sheet.cell(row=cnt,column=8).value = sheet.cell(row=r,column=8).value
##            new_sheet.cell(row=cnt,column=9).value = sheet.cell(row=r,column=9).value
##            new_sheet.cell(row=cnt,column=10).value = sheet.cell(row=r,column=10).value
##            new_sheet.cell(row=cnt,column=11).value = sheet.cell(row=r,column=11).value
##            new_sheet.cell(row=cnt,column=12).value = sheet.cell(row=r,column=12).value
##            new_sheet.cell(row=cnt,column=13).value = sheet.cell(row=r,column=13).value
##            new_sheet.cell(row=cnt,column=14).value = sheet.cell(row=r,column=14).value
##            new_sheet.cell(row=cnt,column=15).value = sheet.cell(row=r,column=15).value
##            new_sheet.cell(row=cnt,column=16).value = sheet.cell(row=r,column=16).value
##            new_sheet.cell(row=cnt,column=17).value = sheet.cell(row=r,column=17).value
##            new_sheet.cell(row=cnt,column=18).value = sheet.cell(row=r,column=18).value
            cnt+=1
    cctv_new.save("trainreg.xlsx")

main()

