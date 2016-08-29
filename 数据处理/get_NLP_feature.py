#-------------------------------------------------------------------------------
# coding=utf8
# Name:        模块1
# Purpose:
#
# Author:      zhx
#
# Created:     26/05/2016
# Copyright:   (c) zhx 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import jieba
import openpyxl
def main():
    positive_senti=[]
    negative_senti=[]
    positive_com =[]
    negative_com =[]
    degree_1 = []
    degree_2 = []
    degree_3 = []
    degree_4 = []
    degree_5 = []
    degree_6 = []
    f = open("pos_sentiment.txt","r")
    lines = f.readlines()#读取全部内容
    cnt = 1
    for line in lines:
        if cnt>2:
            line = line.decode('gbk')
            tmp = line.find(' ')

            positive_senti.append(line[0:tmp])
        cnt+=1
    f.close()

    f = open("neg_sentiment.txt","r")
    lines = f.readlines()#读取全部内容
    cnt = 1
    for line in lines:
        if cnt>2:
            line = line.decode('gbk')
            tmp = line.find(' ')

            negative_senti.append(line[0:tmp])
        cnt+=1
    f.close()

    f = open("pos_com.txt","r")
    lines = f.readlines()#读取全部内容
    cnt = 1
    for line in lines:
        if cnt>2:
            line = line.decode('gbk')
            tmp = line.find(' ')

            positive_com.append(line[0:tmp])
        cnt+=1
    f.close()

    f = open("neg_com.txt","r")
    lines = f.readlines()#读取全部内容
    cnt = 1
    for line in lines:
        if cnt>2:
            line = line.decode('gbk')
            tmp = line.find(' ')

            negative_com.append(line[0:tmp])
        cnt+=1
    f.close()

    f = open("degree.txt","r")
    lines = f.readlines()#读取全部内容
    cnt = 1
    for line in lines:
        line = line.decode('gbk')
        tmp = line.find('\n')
        if cnt>3 and cnt <73:
            degree_6.append(line[0:tmp])
        if cnt>74 and cnt<117:
            degree_5.append(line[0:tmp])
        if cnt>118 and cnt<156:
            degree_4.append(line[0:tmp])
        if cnt>157 and cnt<187:
            degree_3.append(line[0:tmp])
        if cnt>188 and cnt<201:
            degree_2.append(line[0:tmp])
        if cnt>202 and cnt<233:
            degree_1.append(line[0:tmp])
        cnt+=1
    f.close()



    cctv_data = openpyxl.load_workbook("cctv.xlsx")
    cctv_new = openpyxl.Workbook()
    new_sheet = cctv_new.active
    sheet2 = cctv_data["Sheet"]
    for r in xrange(2,4749):
        print r
        content = sheet2.cell(row=r,column=3).value
        if content ==None:
            continue
        seg_list = jieba.cut(content)
        wordsplite = ' '.join(seg_list)
        wordsplite = wordsplite.split(' ')
        num_total = len(wordsplite)
        num_pos = 0
        num_neg = 0
        num_degree_6=0
        num_degree_5=0
        num_degree_4=0
        num_degree_3=0
        num_degree_2=0
        num_degree_1=0
        for w in wordsplite:
            if w in positive_senti or w in positive_com:
                num_pos+=1
            if w in negative_senti or w in negative_com:
                num_neg+=1
            if w in degree_1:
                num_degree_1=1
            if w in degree_2:
                num_degree_2=1
            if w in degree_3:
                num_degree_3=1
            if w in degree_4:
                num_degree_4=1
            if w in degree_5:
                num_degree_5=1
            if w in degree_6:
                num_degree_6=1

        new_sheet.cell(row=r,column=1).value = num_pos
        new_sheet.cell(row=r,column=2).value = round(num_pos/float(num_total) , 3)
        new_sheet.cell(row=r,column=3).value = num_neg
        new_sheet.cell(row=r,column=4).value = round(num_neg/float(num_total) , 3)
        new_sheet.cell(row=r,column=5).value = num_degree_1
        new_sheet.cell(row=r,column=6).value = num_degree_2
        new_sheet.cell(row=r,column=7).value = num_degree_3
        new_sheet.cell(row=r,column=8).value = num_degree_4
        new_sheet.cell(row=r,column=9).value = num_degree_5
        new_sheet.cell(row=r,column=10).value = num_degree_6

    cctv_new.save("train_feature_NLP_title.xlsx")

main()

