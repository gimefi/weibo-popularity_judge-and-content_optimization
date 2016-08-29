#-------------------------------------------------------------------------------
# -*- coding:utf8 -*-
# Name:        模块1
# Purpose:
#
# Author:      zhx
#
# Created:     03/05/2016
# Copyright:   (c) zhx 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import openpyxl

def weekday(data):
    start = data.find('-')
    a = data.find('-',start+1)
    b = data.find(' ')

    m = int(data[start+1:a])
    d = int(data[a+1:b])#.decode('decimal'))

    # all for 2015
    cnt =0
    if m==1:
        cnt+=(d+3)
    if m==2:
        cnt+=(d-1)
    if m==3:
        cnt+=(d-1)
    if m==4:
        cnt+=(d+2)
    if m==5:
        cnt+=(d+4)
    if m==6:
        cnt+=(d)
    if m==7:
        cnt+=(d+2)
    if m==8:
        cnt+=(d+5)
    if m==9:
        cnt+=(d+1)
    if m==10:
        cnt+=(d+3)
    if m==11:
        cnt+=(d-1)
    if m==12:
        cnt+=(d+1)
    return cnt%7


def deal_data(row,column , data):
    if row==1:
        if column ==1:
            return "内容","标题长度","标题内容"
        if column==6:
            return "视频"
        if column==7:
            return "图片"
        if column==8:
            return "艾特"
        if column ==9:
            return "时间","星期"
        if column==10:
            return "话题"
        return data
    if column ==1 :
        #count the length of title
        a = data.find('【')
        b = data.find('】')
        c =b-a-1
        titledata = data[a+1:b]
        tmp = data[0:a]+data[b+1:]
        if c>0:
            return tmp,c,titledata
        else:
            return data,0,''
    if column >=2 and column <=4 :
        return data
    if column ==5:
        # link to web
        a = data.find('O')
        if a==0:
            return 1
        else:
            return 0

    if column==6:
        a = 'L秒拍视频'
        if a in data:
            return 1
        else:
            return 0
    if column==7:
        a = 'picture_count =='
        if a in data:
            return int(data[data.find(a)+17])
        else:
            return 0
    if column==8:
        a = data.find('@')
        if a==0:
            return 1
        else:
            return 0
    if column==9:
        a = data.find(' ')
        week_d = weekday(data)
        return data[a:],week_d
    if column==10:
        # topic or not
        a = data.find('#')
        b = data.find('#',a+1)
        if (b-a)>0:
            return 1
        else:
            return 0
    return
def main():
    cctv_data = openpyxl.load_workbook("cctv15123.xlsx")
    cctv_new = openpyxl.Workbook()
    new_sheet = cctv_new.active
    #print cctv_data.get_sheet_names()
    sheet = cctv_data["Sheet1"]

    for r in xrange(1,1478):
        print r
        for c in xrange(1,11):
            tmp=sheet.cell(row=r,column=c).value
            if c==1:
                res1,res2,res3 = deal_data(r,c,tmp)
                new_sheet.cell(row=r,column=c).value =res1
                new_sheet.cell(row=r,column=c+1).value =res2
                new_sheet.cell(row=r,column=c+2).value =res3
            elif c==9:
                res1,res2 =deal_data(r,c,tmp)
                new_sheet.cell(row=r,column=c+2).value =res1
                new_sheet.cell(row=r,column=c+4).value =res2
            else:
                res = deal_data(r,c,tmp)
                new_sheet.cell(row=r,column=c+2).value =res

    cctv_new.save("cctv_test_formal.xlsx")
main()
