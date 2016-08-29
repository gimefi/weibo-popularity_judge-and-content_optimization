#-------------------------------------------------------------------------------
# coding=utf8
# Name:        模块1
# Purpose:
#
# Author:      zhx
#
# Created:     10/05/2016
# Copyright:   (c) zhx 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import openpyxl
import jieba
threshold = 2140
popular = 0
def main():
    cctv_data = openpyxl.load_workbook("cctv.xlsx")
    cctv_keywords = openpyxl.load_workbook("cctv_keywords.xlsx")
    cctv_new = openpyxl.Workbook()
    new_sheet = cctv_new.active
    #print cctv_data.get_sheet_names()
    sheet1 = cctv_keywords["Sheet"]
    sheet2 = cctv_data["Sheet"]
    words = {}
    for r in xrange(1,36003):
        word = sheet1.cell(row=r,column=1).value
        word_min = sheet1.cell(row=r,column=2).value
        word_max = sheet1.cell(row=r,column=3).value
        word_mean = sheet1.cell(row=r,column=4).value
        words[word] = [word_min,word_max,word_mean]
    for r in xrange(2,4749):
        print r
        content = sheet2.cell(row=r,column=3).value
        time = sheet2.cell(row=r,column=11).value
        like = sheet2.cell(row=r,column=5).value
        repost = sheet2.cell(row=r,column=6).value
        if like == '赞':
            like = '0'
        if repost =='转发':
            repost = '0'
        like_repost = int(like)+int(repost)
        if like_repost>threshold:
            popular =1
        else:
            popular =0
        hour = int(time[1:3])
        minute =int (time[4:])
        time = hour*60 + minute
        new_sheet.cell(row=r,column=10).value = time
        new_sheet.cell(row=r,column=11).value = like_repost
        if content ==None:
            continue
        print r
        seg_list = jieba.cut(content, cut_all = True)
        wordsplite = ' '.join(seg_list)
        wordsplite = wordsplite.split(' ')
        maxlike = 0
        max_word =''
        min_word =''
        mean_word=''
        minlike = 9999999
        tmplist = []
        tmpdic ={}
        for w in wordsplite:
            if words.has_key(w):
                tmpdic[w] =int(words[w][2])
                tmplist.append(int(words[w][2]))
                likes = int(words[w][2])
                if likes<minlike:
                    minlike = likes
                    min_word = w
                if likes>maxlike:
                    maxlike = likes
                    max_word = w
            else:
                continue
        if len(tmplist)!=0:
            tmplist.sort()
            mean = tmplist[int(len(tmplist)/2)]
            for w in tmpdic:
                if tmpdic[w]==mean:
                    mean_word =w

        if min_word!='':
            new_sheet.cell(row=r,column=1).value = words[min_word][0]
            new_sheet.cell(row=r,column=2).value = words[min_word][1]
            new_sheet.cell(row=r,column=3).value = words[min_word][2]
        if max_word!='':
            new_sheet.cell(row=r,column=4).value = words[max_word][0]
            new_sheet.cell(row=r,column=5).value = words[max_word][1]
            new_sheet.cell(row=r,column=6).value = words[max_word][2]
        if mean_word!='':
            new_sheet.cell(row=r,column=7).value = words[mean_word][0]
            new_sheet.cell(row=r,column=8).value = words[mean_word][1]
            new_sheet.cell(row=r,column=9).value = words[mean_word][2]
    cctv_new.save("train_feature_keyword_reg.xlsx")
main()


